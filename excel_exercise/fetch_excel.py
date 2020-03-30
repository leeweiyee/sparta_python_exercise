# Fetch data from the excel file using python module openpyxl

# import module
import openpyxl

# give location of file
path = 'data9.xlsx'

# assign your workbook to a variable
data9_wb = openpyxl.load_workbook(path)
# can use "data9.xlsx" instead of path, but path variable is easier
# e.g. if we need to use it many times and the file is not in the same folder

# storing the data from active sheet into an object called data9
data9_object = data9_wb.active

data9_wb.title = 'Trainees'
data9_title = data9_object.title

# retrieves data in cell A4
cell_value = data9_object.cell(row=1,column=1)

print(cell_value.value)
print(data9_title)

# print total number of rows and columns
count_column = data9_object.max_column
count_row = data9_object.max_row
print("Number of rows: " + str(count_row) + "; Number of columns: " + str(count_column))

# rename sheet
# go to the python file you create the excel file and rename there

# change value using python
cell_value.value = "Name"
print(cell_value.value)

data9_wb.save(path)