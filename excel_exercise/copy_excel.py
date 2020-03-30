# Task: copy data from data9 to a new file
import openpyxl
import xlsxwriter

# create destination excel file
data9copy = xlsxwriter.Workbook("data9copy.xlsx") #creating a workbook using the built in method and assigning a name to the file
worksheet = data9copy.add_worksheet("Trainees")
data9copy.close()

# open source excel file
path = "data9.xlsx" # give location of the file
data9_wb = openpyxl.load_workbook(path) # create an object
data9_ws = data9_wb.worksheets[0] # gets the data from the file

# open destination excel file
new_path = "data9copy.xlsx"
data9copy_wb = openpyxl.load_workbook(new_path)
data9copy_ws =data9copy_wb.worksheets[0]

number_of_rows = data9_ws.max_row
number_of_columns = data9_ws.max_column

# copy  cell values from source to destination
for i in range(1, number_of_rows + 1):
    for j in range(1, number_of_columns + 1):
        # fetch cell value from source excel file
        cell_content = data9_ws.cell(row = i, column = j)
        # writing fetched value to destination excel file
        data9copy_ws.cell(row=i, column=j).value = cell_content.value

# saving the destination excel file
data9copy_wb.save(str(new_path))